from openpyxl import load_workbook
from collections import OrderedDict
from datetime import datetime


# First we get the date from the original sheet

# Load the excel file
wb = load_workbook("/users/yichengzhang/desktop/ControlHome1.xlsx")
print(wb.get_sheet_names()[0])

# Get the first sheet(Date/Time)
ws_raw = wb.get_sheet_by_name(wb.get_sheet_names()[0])

# Get the first column(Value)
date_times = ws_raw.columns[0]

# Get the second column
values = ws_raw.columns[1]

# Ordered dictionary that contains consumption result,
# key is the date(yyyy-mm-dd),
# value is also an ordered dictionary,
# of which the key is the time(hh:mm:ss)
# the value is the consumption value of that time
consumption = OrderedDict()

# Iterate through all the times
for i in range(1, len(date_times)):
    # Get the ith cell of Date/Time
    date_time = date_times[i]

    # Split the string by space into a list,
    # first element is date, second element is time

    # date_object = datetime.strptime((str)(date_time.value), '%Y-%m-%d %H:%M:%S')
    # date_string = date_object.strftime('%Y-%m-%d %H:%M:%S')
    date_time_list = (str)(date_time.value).split()
    # date = "yyyy-mm-dd"
    date = date_time_list[0]
    # time = "hh:mm:ss"
    time = date_time_list[1]

    # If our ordered dictionary contains
    # this date key, we get its corresponding value
    # which is an ordered dictionary
    if date in consumption:
        time_dict = consumption[date]
    # If not, we create a new ordered dictionary
    else:
        time_dict = OrderedDict()

    # Then we put the key value pair in the inner dict
    time_dict[time] = values[i].value

    # Then put the key value pair to the consumption dict
    consumption[date] = time_dict

print(consumption.keys())

print(consumption[consumption.keys()[0]])


# Now we put our data into the new sheet

# Create a new empty sheet
ws_consumption = wb.create_sheet()

# Set the sheet title
ws_consumption.title = "Consumption"

# Write date to the sheet
ws_consumption["A1"] = "Plot Name"
ws_consumption["C1"] = "Day Type"
ws_consumption["E1"] = "Occupants Info"

ws_consumption["A2"] = "House Type"
ws_consumption["C2"] = "Month"


# Iterate through dict and put date in the sheet
for col in range(0, len(consumption), 2):
    date = consumption.keys()[col]
    print(date)
    ordered_dict = consumption.get(date)
    _ = ws_consumption.cell(row=1 + 2, column=col + 1, value="Time")
    _ = "Time"
    _ = ws_consumption.cell(row=1 + 2, column=col + 2, value="Value")
    _ = "Value"

    for row in range(0, len(ordered_dict)):
        _ = date + " " + time
        time = ordered_dict.keys()[row]
        _ = ws_consumption.cell(row=row + 1 + 3, column=col + 1, value=date + " " + time)

        _ = ws_consumption.cell(row=row + 1 + 3, column=col + 2, value=ordered_dict.get(time))


# Save the sheet
wb.save("/users/yichengzhang/desktop/ControlHome1.xlsx")
